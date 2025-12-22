"""
Service for Excel import/export operations
"""
import pandas as pd
import sqlite3
from datetime import datetime
from typing import List, Dict, Any, Optional
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from backend.config import settings
from backend.models.wbs import WBSCreate
from backend.services.wbs_service import WBSService


class ExcelService:
    """Service for Excel import/export"""

    def __init__(self):
        self.db_path = str(settings.database_path)
        self.wbs_service = WBSService()

    def _parse_date(self, date_str: Any) -> Optional[str]:
        """Parse date from various formats to YYYY-MM-DD"""
        if pd.isna(date_str) or date_str is None or date_str == '':
            return None

        try:
            # If already a datetime object
            if isinstance(date_str, datetime):
                return date_str.strftime('%Y-%m-%d')

            # If string, try to parse
            date_str = str(date_str).strip()
            if not date_str:
                return None

            # Try different formats (prioritize yyyy/mm/dd format)
            for fmt in ['%Y/%m/%d', '%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y']:
                try:
                    dt = datetime.strptime(date_str, fmt)
                    return dt.strftime('%Y-%m-%d')
                except ValueError:
                    continue

            # Try pandas to_datetime as fallback
            dt = pd.to_datetime(date_str)
            return dt.strftime('%Y-%m-%d')

        except Exception:
            return None

    def import_wbs_from_excel(self, file_path: str, project_id: str) -> Dict[str, Any]:
        """
        Import WBS items from Excel file

        Expected columns (Chinese):
        - 項目 (WBS ID)
        - 父項目 (Parent WBS ID)
        - 任務說明 (Task Name)
        - 單位 (Owner Unit)
        - 類別 (Category)
        - 預計開始 (原始) (Original Planned Start)
        - 預計結束 (原始) (Original Planned End)
        - 預計開始 (調整) (Revised Planned Start)
        - 預計結束 (調整) (Revised Planned End)
        - 開始日期 (Actual Start)
        - 結束日期 (Actual End)
        - 工作天數 (Work Days)
        - 實際完成進度 (Actual Progress)
        - 狀態 (Status)
        - 備註說明 (Notes)
        """
        try:
            # Read Excel file
            df = pd.read_excel(file_path)

            # Column mapping (Chinese to English)
            column_map = {
                '項目': 'wbs_id',
                '父項目': 'parent_id',
                '任務說明': 'task_name',
                '單位': 'owner_unit',
                '類別': 'category',
                '預計開始 (原始)': 'original_planned_start',
                '預計結束 (原始)': 'original_planned_end',
                '預計開始 (調整)': 'revised_planned_start',
                '預計結束 (調整)': 'revised_planned_end',
                '開始日期': 'actual_start_date',
                '結束日期': 'actual_end_date',
                '工作天數': 'work_days',
                '實際完成進度': 'actual_progress',
                '狀態': 'status',
                '備註說明': 'notes',
                '內部安排': 'is_internal',
            }

            # Only rename columns that exist in the DataFrame
            existing_column_map = {k: v for k, v in column_map.items() if k in df.columns}
            df = df.rename(columns=existing_column_map)

            # Required columns check
            required_columns = ['wbs_id', 'task_name']
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                return {
                    'success': False,
                    'error': f'Missing required columns: {", ".join(missing_columns)}',
                    'imported': 0,
                    'failed': len(df)
                }

            # Import WBS items
            imported = []
            failed = []

            for idx, row in df.iterrows():
                try:
                    # Skip if WBS ID is empty
                    if pd.isna(row.get('wbs_id')) or str(row.get('wbs_id')).strip() == '':
                        continue

                    # Prepare WBS data
                    # Handle parent_id - convert to string and clean up
                    parent_id_value = None
                    if pd.notna(row.get('parent_id')):
                        parent_str = str(row.get('parent_id')).strip()
                        # Only remove .0 suffix if it's purely numeric (like 1.0 -> 1, 2.0 -> 2)
                        # But keep WBS IDs like 2.2, 1.1.3 intact
                        if parent_str and parent_str != 'nan':
                            # Check if it ends with .0 and has only one decimal point (purely numeric)
                            if parent_str.endswith('.0'):
                                # Check if removing .0 leaves a valid integer
                                base = parent_str[:-2]  # Remove .0
                                if base.isdigit():
                                    parent_str = base
                            if parent_str:
                                parent_id_value = parent_str

                    # Handle is_internal - convert various representations to boolean
                    # Default to False if column doesn't exist in the Excel file
                    is_internal_value = False
                    if 'is_internal' in df.columns:
                        if pd.notna(row.get('is_internal')):
                            val = str(row.get('is_internal')).strip().lower()
                            # Accept: 'yes', 'y', 'true', '1', '是', 'v', '✓'
                            is_internal_value = val in ['yes', 'y', 'true', '1', '是', 'v', '✓', 'x']

                    wbs_data = {
                        'project_id': project_id,
                        'wbs_id': str(row['wbs_id']).strip(),
                        'parent_id': parent_id_value,
                        'task_name': str(row.get('task_name', '')).strip(),
                        'category': str(row.get('category', 'Task')).strip() or 'Task',
                        'owner_unit': str(row.get('owner_unit', '')).strip() if pd.notna(row.get('owner_unit')) else None,
                        'original_planned_start': self._parse_date(row.get('original_planned_start')),
                        'original_planned_end': self._parse_date(row.get('original_planned_end')),
                        'revised_planned_start': self._parse_date(row.get('revised_planned_start')),
                        'revised_planned_end': self._parse_date(row.get('revised_planned_end')),
                        'actual_start_date': self._parse_date(row.get('actual_start_date')),
                        'actual_end_date': self._parse_date(row.get('actual_end_date')),
                        'work_days': int(row.get('work_days', 0)) if pd.notna(row.get('work_days')) else None,
                        'actual_progress': int(row.get('actual_progress', 0)) if pd.notna(row.get('actual_progress')) else 0,
                        'status': str(row.get('status', '未開始')).strip() or '未開始',
                        'notes': str(row.get('notes', '')).strip() if pd.notna(row.get('notes')) else None,
                        'is_internal': is_internal_value,
                    }

                    # Create WBS item
                    wbs_create = WBSCreate(**wbs_data)
                    result = self.wbs_service.create_wbs(wbs_create)

                    imported.append({
                        'row': idx + 2,  # Excel row number (1-indexed + header)
                        'wbs_id': wbs_data['wbs_id'],
                        'task_name': wbs_data['task_name']
                    })

                except Exception as e:
                    failed.append({
                        'row': idx + 2,
                        'wbs_id': str(row.get('wbs_id', 'N/A')),
                        'error': str(e)
                    })

            return {
                'success': len(imported) > 0,  # 至少要有一筆成功才算成功
                'imported': len(imported),
                'failed': len(failed),
                'imported_items': imported,
                'failed_items': failed
            }

        except Exception as e:
            return {
                'success': False,
                'error': str(e),
                'imported': 0,
                'failed': 0
            }

    def export_wbs_to_excel(self, project_id: str, output_path: str) -> Dict[str, Any]:
        """
        Export WBS items to Excel file

        Returns styled Excel file with all WBS data
        """
        try:
            # Get WBS data
            conn = sqlite3.connect(self.db_path)
            conn.row_factory = sqlite3.Row

            query = """
                SELECT
                    wbs_id,
                    parent_id,
                    task_name,
                    owner_unit,
                    category,
                    original_planned_start,
                    original_planned_end,
                    revised_planned_start,
                    revised_planned_end,
                    actual_start_date,
                    actual_end_date,
                    work_days,
                    actual_progress,
                    estimated_progress,
                    progress_variance,
                    status,
                    notes,
                    is_internal,
                    is_overdue
                FROM tracking_items
                WHERE project_id = ? AND item_type = 'WBS'
                ORDER BY wbs_id
            """

            df = pd.read_sql_query(query, conn, params=(project_id,))
            conn.close()

            if df.empty:
                return {
                    'success': False,
                    'error': 'No WBS items found for this project',
                    'exported': 0
                }

            # Rename columns to Chinese
            df = df.rename(columns={
                'wbs_id': '項目',
                'parent_id': '父項目',
                'task_name': '任務說明',
                'owner_unit': '單位',
                'category': '類別',
                'original_planned_start': '預計開始 (原始)',
                'original_planned_end': '預計結束 (原始)',
                'revised_planned_start': '預計開始 (調整)',
                'revised_planned_end': '預計結束 (調整)',
                'actual_start_date': '開始日期',
                'actual_end_date': '結束日期',
                'work_days': '工作天數',
                'actual_progress': '實際完成進度',
                'estimated_progress': '預估完成進度',
                'progress_variance': '進度偏差',
                'status': '狀態',
                'notes': '備註說明',
                'is_internal': '內部安排',
                'is_overdue': '逾期'
            })

            # Convert boolean to Chinese
            df['逾期'] = df['逾期'].apply(lambda x: '是' if x == 1 else '否')
            df['內部安排'] = df['內部安排'].apply(lambda x: 'V' if x == 1 else '')

            # Create Excel writer
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='WBS', index=False)

                # Get workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['WBS']

                # Style header
                header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                header_font = Font(bold=True, color='FFFFFF')
                border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = border

                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width

                # Apply borders to all cells
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row,
                                               min_col=1, max_col=worksheet.max_column):
                    for cell in row:
                        cell.border = border

                # Highlight overdue items
                overdue_fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
                overdue_col_idx = list(df.columns).index('逾期') + 1

                for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), start=2):
                    if row[overdue_col_idx - 1].value == '是':
                        for cell in row:
                            cell.fill = overdue_fill

            return {
                'success': True,
                'exported': len(df),
                'file_path': output_path
            }

        except Exception as e:
            return {
                'success': False,
                'error': str(e),
                'exported': 0
            }

    def create_wbs_template(self, output_path: str) -> Dict[str, Any]:
        """
        Create WBS import template Excel file
        """
        try:
            # Define template columns
            columns = [
                '項目',
                '父項目',
                '任務說明',
                '單位',
                '類別',
                '預計開始 (原始)',
                '預計結束 (原始)',
                '預計開始 (調整)',
                '預計結束 (調整)',
                '開始日期',
                '結束日期',
                '工作天數',
                '實際完成進度',
                '狀態',
                '內部安排',
                '備註說明'
            ]

            # Create sample data
            sample_data = [
                ['1', '', '專案啟動', '專案經理', 'Milestone', '01/01/2024', '01/01/2024', '', '', '', '', '', 100, '已完成', '', '頂層項目範例'],
                ['1.1', '1', '需求分析', '開發部', 'Task', '01/02/2024', '01/15/2024', '', '', '01/02/2024', '01/14/2024', 10, 100, '已完成', '', '子項目範例'],
                ['1.2', '1', '系統設計', 'AAA/BBB', 'Task', '01/16/2024', '01/31/2024', '', '', '01/16/2024', '', 12, 60, '進行中', 'V', '子項目範例（內部安排）'],
                ['2', '', '開發階段', '開發部', 'Milestone', '02/01/2024', '03/31/2024', '', '', '', '', '', 0, '未開始', '', '頂層項目範例'],
            ]

            df = pd.DataFrame(sample_data, columns=columns)

            # Create Excel file
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='WBS範本', index=False)

                workbook = writer.book
                worksheet = writer.sheets['WBS範本']

                # Style header
                header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                header_font = Font(bold=True, color='FFFFFF')

                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                # Auto-adjust columns
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 30)
                    worksheet.column_dimensions[column_letter].width = adjusted_width

            return {
                'success': True,
                'file_path': output_path,
                'message': 'Template created successfully'
            }

        except Exception as e:
            return {
                'success': False,
                'error': str(e)
            }

    def export_pending_to_excel(self, project_id: str, output_path: str) -> Dict[str, Any]:
        """
        Export Pending items to Excel file

        Returns styled Excel file with all Pending data
        """
        try:
            # Get Pending data
            conn = sqlite3.connect(self.db_path)
            conn.row_factory = sqlite3.Row

            query = """
                SELECT
                    pending_id,
                    task_date,
                    source_type,
                    contact_info,
                    description,
                    planned_start_date,
                    expected_completion_date,
                    actual_completion_date,
                    is_replied,
                    handling_notes,
                    related_wbs,
                    related_action_item,
                    status,
                    priority
                FROM pending_items
                WHERE project_id = ?
                ORDER BY task_date DESC
            """

            df = pd.read_sql_query(query, conn, params=(project_id,))
            conn.close()

            if df.empty:
                return {
                    'success': False,
                    'error': 'No pending items found for this project',
                    'exported': 0
                }

            # Rename columns to Chinese
            df = df.rename(columns={
                'pending_id': '編號',
                'task_date': '任務日期',
                'source_type': '來源類型',
                'contact_info': '聯絡資訊',
                'description': '說明',
                'planned_start_date': '預計開始日期',
                'expected_completion_date': '預計完成日期',
                'actual_completion_date': '實際完成日期',
                'is_replied': '已回覆',
                'handling_notes': '處理備註',
                'related_wbs': '相關 WBS',
                'related_action_item': '相關行動項目',
                'status': '狀態',
                'priority': '優先級'
            })

            # Convert boolean to Chinese
            df['已回覆'] = df['已回覆'].apply(lambda x: '是' if x == 1 else '否')

            # Create Excel writer
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='待辦事項', index=False)

                # Get workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['待辦事項']

                # Style header
                header_fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
                header_font = Font(bold=True, color='FFFFFF')
                border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = border

                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width

                # Apply borders to all cells
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row,
                                               min_col=1, max_col=worksheet.max_column):
                    for cell in row:
                        cell.border = border

            return {
                'success': True,
                'exported': len(df),
                'file_path': output_path
            }

        except Exception as e:
            return {
                'success': False,
                'error': str(e),
                'exported': 0
            }

    def export_issues_to_excel(self, project_id: str, output_path: str) -> Dict[str, Any]:
        """
        Export Issues to Excel file

        Returns styled Excel file with all Issue data
        """
        try:
            # Get Issues data
            conn = sqlite3.connect(self.db_path)
            conn.row_factory = sqlite3.Row

            query = """
                SELECT
                    issue_number,
                    issue_title,
                    issue_description,
                    issue_type,
                    issue_category,
                    severity,
                    priority,
                    reported_by,
                    reported_date,
                    assigned_to,
                    owner_type,
                    affected_wbs,
                    impact_description,
                    estimated_impact_days,
                    status,
                    resolution,
                    root_cause,
                    target_resolution_date,
                    actual_resolution_date,
                    closed_date,
                    is_escalated,
                    escalation_level,
                    escalation_date,
                    escalation_reason
                FROM issue_tracking
                WHERE project_id = ?
                ORDER BY issue_number
            """

            df = pd.read_sql_query(query, conn, params=(project_id,))
            conn.close()

            if df.empty:
                return {
                    'success': False,
                    'error': 'No issues found for this project',
                    'exported': 0
                }

            # Rename columns to Chinese
            df = df.rename(columns={
                'issue_number': '問題編號',
                'issue_title': '問題標題',
                'issue_description': '問題說明',
                'issue_type': '問題類型',
                'issue_category': '問題分類',
                'severity': '嚴重性',
                'priority': '優先級',
                'reported_by': '回報人',
                'reported_date': '回報日期',
                'assigned_to': '指派給',
                'owner_type': '負責類型',
                'affected_wbs': '受影響 WBS',
                'impact_description': '影響說明',
                'estimated_impact_days': '預估影響天數',
                'status': '狀態',
                'resolution': '解決方案',
                'root_cause': '根本原因',
                'target_resolution_date': '目標解決日期',
                'actual_resolution_date': '實際解決日期',
                'closed_date': '關閉日期',
                'is_escalated': '已升級',
                'escalation_level': '升級層級',
                'escalation_date': '升級日期',
                'escalation_reason': '升級原因'
            })

            # Convert boolean to Chinese
            df['已升級'] = df['已升級'].apply(lambda x: '是' if x == 1 else '否')

            # Create Excel writer
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='問題追蹤', index=False)

                # Get workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['問題追蹤']

                # Style header
                header_fill = PatternFill(start_color='E74C3C', end_color='E74C3C', fill_type='solid')
                header_font = Font(bold=True, color='FFFFFF')
                border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = border

                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width

                # Apply borders to all cells
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row,
                                               min_col=1, max_col=worksheet.max_column):
                    for cell in row:
                        cell.border = border

                # Highlight escalated items
                escalated_fill = PatternFill(start_color='FFF4E6', end_color='FFF4E6', fill_type='solid')
                escalated_col_idx = list(df.columns).index('已升級') + 1

                for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), start=2):
                    if row[escalated_col_idx - 1].value == '是':
                        for cell in row:
                            cell.fill = escalated_fill

            return {
                'success': True,
                'exported': len(df),
                'file_path': output_path
            }

        except Exception as e:
            return {
                'success': False,
                'error': str(e),
                'exported': 0
            }
